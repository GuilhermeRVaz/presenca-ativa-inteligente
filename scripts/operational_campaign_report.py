from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

load_dotenv(ROOT_DIR / ".env")

from app.core.config import settings
from app.infrastructure.supabase.repositories import SupabaseRepository


SCHEMA = "busca_ativa_v2"


@dataclass(frozen=True)
class ConversationGroup:
    sender_jid: str
    push_name: str
    resolved: bool
    response_count: int
    first_received_at: str
    last_received_at: str
    consolidated_text: str
    category: str
    confidence_values: str
    suggested_student: str
    suggested_score: float
    suggested_note: str


def normalize_text(value: str | None) -> str:
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_text(value: str | None) -> str:
    text = re.sub(r"<Mensagem editada>", "", str(value or ""), flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def category_for_text(text: str) -> str:
    norm = normalize_text(text)
    if not norm or norm == "SEM JUSTIFICATIVA":
        return "SEM_JUSTIFICATIVA_TEXTUAL"
    if "MIDIA OCULTA" in norm and len(norm) <= 40:
        return "MIDIA_SEM_TEXTO"

    categories = [
        (
            "SAUDE",
            [
                "VIROSE",
                "FEBRE",
                "GRIPE",
                "GRIPAD",
                "INFLUENZA",
                "DOENTE",
                "DOR",
                "GARGANTA",
                "ASMA",
                "ALERG",
                "INDISPOST",
                "DENT",
                "MEDICO",
                "REMEDIO",
                "SONO",
                "CRISE",
                "VACINA",
                "HOSPITAL",
                "CAMA",
                "NADA BEM",
            ],
        ),
        (
            "FAMILIAR_LOGISTICA",
            [
                "NAO CONSEGUI",
                "NAO DEU",
                "LEVAR",
                "BUSCA",
                "COMPROMISSO",
                "VIAGEM",
                "VIAJAR",
                "PERDEU A HR",
                "PERDEU A HORA",
                "ATRAS",
            ],
        ),
        ("ERRO_CHAMADA_PRESENCA", ["FOI PARA", "FOI PRA", "ESTAVA NA ESCOLA", "NAO FALTOU", "INCORRETO"]),
        ("AGRADECIMENTO_SEM_MOTIVO", ["OBRIGAD", "AGRADECO", "AGRADEÇO", "BOM DIA", "BOA TARDE"]),
        ("RECADO_OUTRO_ASSUNTO", ["RECADO", "PASSAR UM RECADO"]),
    ]
    for category, terms in categories:
        if any(term in norm for term in terms):
            return category
    return "OUTRA_JUSTIFICATIVA"


def student_name_score(student_name: str, text: str) -> float:
    student_norm = normalize_text(student_name)
    text_norm = normalize_text(text)
    if not student_norm or not text_norm:
        return 0.0
    if student_norm in text_norm:
        return 1.0

    tokens = student_norm.split()
    first_token_hit = bool(tokens and tokens[0] in text_norm)
    token_hits = sum(1 for token in tokens if len(token) > 2 and token in text_norm)
    token_score = token_hits / max(len(tokens), 1)
    ratio = SequenceMatcher(None, student_norm, text_norm[: max(len(student_norm) * 2, 80)]).ratio()
    score = max(token_score, ratio)
    if not first_token_hit:
        score = min(score, 0.49)
    return score


def fetch_all(query: Any, *, page_size: int = 1000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    start = 0
    while True:
        page = query.range(start, start + page_size - 1).execute().data or []
        rows.extend(page)
        if len(page) < page_size:
            return rows
        start += page_size


def build_client() -> Any:
    if not settings.supabase_url or not settings.supabase_key:
        raise RuntimeError("SUPABASE_URL e SUPABASE_KEY precisam estar configurados no .env")
    return SupabaseRepository().client.schema(SCHEMA)


def load_campaign(client: Any, campaign_id: str, school_id: str) -> dict[str, Any]:
    rows = (
        client.table("campaigns")
        .select("id,name,status,absence_days,created_at,total_sent,total_replied")
        .eq("id", campaign_id)
        .eq("school_id", school_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise RuntimeError(f"Campanha nao encontrada: {campaign_id}")
    return rows[0]


def load_messages(client: Any, campaign_id: str, school_id: str) -> list[dict[str, Any]]:
    return fetch_all(
        client.table("messages")
        .select(
            "id,status,tracking_ref,evolution_msg_id,wa_jid,template_id,body_preview,"
            "sent_at,delivered_at,read_at,replied_at,created_at,"
            "student_id,guardian_id,"
            "students(id,ra,name,class_name),"
            "guardians(id,name,phone_e164,wa_jid)"
        )
        .eq("campaign_id", campaign_id)
        .eq("school_id", school_id)
        .order("created_at")
    )


def load_responses(client: Any, campaign_id: str, school_id: str) -> list[dict[str, Any]]:
    return fetch_all(
        client.table("responses")
        .select(
            "id,message_id,guardian_id,campaign_id,student_id,raw_message_id,sender_jid,"
            "body,identity_confidence,classified,reason,ai_confidence,is_ack,needs_review,"
            "received_at,created_at"
        )
        .eq("campaign_id", campaign_id)
        .eq("school_id", school_id)
        .order("received_at")
    )


def load_sessions_for_senders(client: Any, school_id: str, senders: set[str]) -> dict[str, dict[str, Any]]:
    sessions: dict[str, dict[str, Any]] = {}
    for sender in sorted(senders):
        rows = (
            client.table("conversation_sessions")
            .select("*")
            .eq("school_id", school_id)
            .eq("sender_jid", sender)
            .limit(1)
            .execute()
            .data
            or []
        )
        if rows:
            sessions[sender] = rows[0]
    return sessions


def suggest_student(messages: list[dict[str, Any]], text: str) -> tuple[str, float, str]:
    scored: list[tuple[float, str]] = []
    for message in messages:
        student = message.get("students") or {}
        name = str(student.get("name") or "")
        score = student_name_score(name, text)
        scored.append((score, name))
    scored.sort(reverse=True, key=lambda item: item[0])
    if not scored:
        return "", 0.0, ""
    top_score, top_name = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    delta = top_score - second_score
    if top_score >= 0.58 and delta >= 0.15:
        return top_name, top_score, "SUGESTAO_TEXTUAL_NAO_CONFIRMADA"
    return "", top_score, "SEM_SUGESTAO_SEGURA"


def build_conversation_groups(
    *,
    client: Any,
    school_id: str,
    messages: list[dict[str, Any]],
    responses: list[dict[str, Any]],
) -> list[ConversationGroup]:
    by_sender: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for response in responses:
        by_sender[str(response.get("sender_jid") or "")].append(response)

    sessions = load_sessions_for_senders(client, school_id, set(by_sender))
    groups: list[ConversationGroup] = []
    for sender, rows in sorted(by_sender.items(), key=lambda item: item[1][0].get("received_at") or ""):
        session = sessions.get(sender, {})
        texts = [clean_text(row.get("body")) for row in rows]
        non_empty_texts = [text if text else "[MIDIA_OU_MENSAGEM_SEM_TEXTO]" for text in texts]
        consolidated = " | ".join(non_empty_texts).strip() or "SEM JUSTIFICATIVA"
        suggestion_text = f"{session.get('push_name') or ''} | {consolidated}"
        suggested_student, suggested_score, suggested_note = suggest_student(messages, suggestion_text)
        groups.append(
            ConversationGroup(
                sender_jid=sender,
                push_name=str(session.get("push_name") or ""),
                resolved=bool(session.get("resolved")),
                response_count=len(rows),
                first_received_at=str(rows[0].get("received_at") or ""),
                last_received_at=str(rows[-1].get("received_at") or ""),
                consolidated_text=consolidated,
                category=category_for_text(consolidated),
                confidence_values=", ".join(sorted({str(row.get("identity_confidence") or "") for row in rows})),
                suggested_student=suggested_student,
                suggested_score=round(suggested_score, 3),
                suggested_note=suggested_note,
            )
        )
    return groups


def render_markdown(
    *,
    campaign: dict[str, Any],
    messages: list[dict[str, Any]],
    responses: list[dict[str, Any]],
    groups: list[ConversationGroup],
) -> str:
    message_status = Counter(str(row.get("status") or "NULL") for row in messages)
    response_confidence = Counter(str(row.get("identity_confidence") or "NULL") for row in responses)
    group_categories = Counter(group.category for group in groups)
    identified_responses = [
        row for row in responses if row.get("student_id") and row.get("guardian_id")
    ]

    lines = [
        "# Relatorio Operacional Automatico v1",
        "",
        f"- Campanha: {campaign.get('name')}",
        f"- Campaign ID: {campaign.get('id')}",
        f"- Dias de falta: {campaign.get('absence_days')}",
        f"- Status: {campaign.get('status')}",
        f"- Criada em: {campaign.get('created_at')}",
        "",
        "## Metricas",
        "",
        "| Metrica | Valor |",
        "| --- | ---: |",
        f"| Alunos/destinatarios enfileirados | {len(messages)} |",
        f"| Responses no banco | {len(responses)} |",
        f"| Interlocutores/sessoes com resposta | {len(groups)} |",
        f"| Responses identificadas com aluno e responsavel | {len(identified_responses)} |",
        f"| Responses sem identidade confirmada | {len(responses) - len(identified_responses)} |",
        "",
        "Status de envio:",
        *[f"- {status}: {total}" for status, total in sorted(message_status.items())],
        "",
        "Confianca das responses:",
        *[f"- {confidence}: {total}" for confidence, total in sorted(response_confidence.items())],
        "",
        "Categorias preliminares por interlocutor:",
        *[f"- {category}: {total}" for category, total in group_categories.most_common()],
        "",
        "## Bloco 1 - Destinatarios da campanha",
        "",
        "| RA | Aluno | Turma | Responsavel | Telefone | Status envio |",
        "| --- | --- | --- | --- | --- | --- |",
    ]

    for message in messages:
        student = message.get("students") or {}
        guardian = message.get("guardians") or {}
        lines.append(
            "| "
            + " | ".join(
                [
                    str(student.get("ra") or ""),
                    str(student.get("name") or ""),
                    str(student.get("class_name") or ""),
                    str(guardian.get("name") or ""),
                    str(guardian.get("phone_e164") or ""),
                    str(message.get("status") or ""),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Bloco 2 - Respostas agrupadas por interlocutor",
            "",
            "Estas respostas estao vinculadas a campanha e a uma sessao, mas nao necessariamente a um aluno/responsavel confirmado.",
            "",
        ]
    )

    for idx, group in enumerate(groups, 1):
        lines.extend(
            [
                f"### Interlocutor {idx:02d}",
                "",
                f"- Sender JID: `{group.sender_jid}`",
                f"- Push name: {group.push_name or 'NAO_INFORMADO'}",
                f"- Mensagens: {group.response_count}",
                f"- Primeira resposta: {group.first_received_at}",
                f"- Ultima resposta: {group.last_received_at}",
                f"- Sessao resolvida: {'sim' if group.resolved else 'nao'}",
                f"- Confianca atual: {group.confidence_values or 'NAO_INFORMADA'}",
                f"- Categoria preliminar: {group.category}",
                f"- Sugestao textual: {group.suggested_student or 'SEM_SUGESTAO_SEGURA'} ({group.suggested_score:.3f})",
                "",
                "Texto consolidado:",
                "",
                f"> {group.consolidated_text}",
                "",
            ]
        )

    lines.extend(
        [
            "## Leitura operacional",
            "",
            "- Este relatorio nao inventa vinculo aluno/responsavel quando a identidade ainda esta UNRESOLVED.",
            "- O Bloco 1 e factual: campanha, alunos, responsaveis cadastrados e status de envio.",
            "- O Bloco 2 e conversacional: respostas reais agrupadas por sessao/sender_jid.",
            "- Sugestoes textuais sao apenas apoio de auditoria; nao atualizam o banco.",
        ]
    )
    return "\n".join(lines) + "\n"


def write_csvs(out_dir: Path, stem: str, messages: list[dict[str, Any]], groups: list[ConversationGroup]) -> tuple[Path, Path]:
    recipients_path = out_dir / f"{stem}_destinatarios.csv"
    groups_path = out_dir / f"{stem}_interlocutores.csv"

    with recipients_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = ["ra", "aluno", "turma", "responsavel", "telefone", "status_envio", "sent_at"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for message in messages:
            student = message.get("students") or {}
            guardian = message.get("guardians") or {}
            writer.writerow(
                {
                    "ra": student.get("ra") or "",
                    "aluno": student.get("name") or "",
                    "turma": student.get("class_name") or "",
                    "responsavel": guardian.get("name") or "",
                    "telefone": guardian.get("phone_e164") or "",
                    "status_envio": message.get("status") or "",
                    "sent_at": message.get("sent_at") or "",
                }
            )

    with groups_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = [
            "sender_jid",
            "push_name",
            "response_count",
            "first_received_at",
            "last_received_at",
            "category",
            "confidence_values",
            "suggested_student",
            "suggested_score",
            "suggested_note",
            "consolidated_text",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for group in groups:
            writer.writerow({field: getattr(group, field) for field in fieldnames})

    return recipients_path, groups_path


def write_xlsx(out_dir: Path, stem: str, messages: list[dict[str, Any]], groups: list[ConversationGroup]) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    path = out_dir / f"{stem}.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Destinatarios"
    headers = ["RA", "Aluno", "Turma", "Responsavel", "Telefone", "Status envio", "Enviado em"]
    ws.append(headers)
    for message in messages:
        student = message.get("students") or {}
        guardian = message.get("guardians") or {}
        ws.append(
            [
                student.get("ra") or "",
                student.get("name") or "",
                student.get("class_name") or "",
                guardian.get("name") or "",
                guardian.get("phone_e164") or "",
                message.get("status") or "",
                message.get("sent_at") or "",
            ]
        )

    ws2 = wb.create_sheet("Interlocutores")
    ws2.append(
        [
            "Sender JID",
            "Push name",
            "Mensagens",
            "Primeira resposta",
            "Ultima resposta",
            "Categoria",
            "Confianca",
            "Sugestao textual",
            "Score sugestao",
            "Texto consolidado",
        ]
    )
    for group in groups:
        ws2.append(
            [
                group.sender_jid,
                group.push_name,
                group.response_count,
                group.first_received_at,
                group.last_received_at,
                group.category,
                group.confidence_values,
                group.suggested_student,
                group.suggested_score,
                group.consolidated_text,
            ]
        )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for col_idx in range(1, sheet.max_column + 1):
            max_len = 10
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx):
                max_len = max(max_len, min(len(str(row[0].value or "")), 80))
            sheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(path)
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera relatorio operacional automatico v1 de uma campanha.")
    parser.add_argument("--campaign-id", required=True)
    parser.add_argument("--school-id", default=settings.default_school_id)
    parser.add_argument("--out-dir", type=Path, default=ROOT_DIR / "relatorios" / "campanhas")
    args = parser.parse_args()

    if not args.school_id:
        raise SystemExit("DEFAULT_SCHOOL_ID nao configurado e --school-id nao informado.")

    client = build_client()
    campaign = load_campaign(client, args.campaign_id, args.school_id)
    messages = load_messages(client, args.campaign_id, args.school_id)
    responses = load_responses(client, args.campaign_id, args.school_id)
    groups = build_conversation_groups(
        client=client,
        school_id=args.school_id,
        messages=messages,
        responses=responses,
    )

    args.out_dir.mkdir(parents=True, exist_ok=True)
    campaign_date = str(campaign.get("absence_days") or datetime.now().date()).replace("/", "_").replace("-", "_")
    stem = f"relatorio_operacional_{campaign_date}_{args.campaign_id[:8]}"
    markdown = render_markdown(campaign=campaign, messages=messages, responses=responses, groups=groups)
    md_path = args.out_dir / f"{stem}.md"
    md_path.write_text(markdown, encoding="utf-8")
    recipients_csv, groups_csv = write_csvs(args.out_dir, stem, messages, groups)
    xlsx_path = write_xlsx(args.out_dir, stem, messages, groups)

    try:
        print(markdown)
    except UnicodeEncodeError:
        print(markdown.encode('ascii', errors='replace').decode('ascii'))
    print(f"Relatorio Markdown: {md_path}")
    print(f"CSV destinatarios: {recipients_csv}")
    print(f"CSV interlocutores: {groups_csv}")
    if xlsx_path:
        print(f"XLSX: {xlsx_path}")
    else:
        print("XLSX nao gerado: openpyxl nao esta instalado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
