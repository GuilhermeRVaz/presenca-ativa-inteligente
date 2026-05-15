"""
RELATORIO DE EXPLORACAO - Busca Ativa Inteligente
Data: 2026-04-29
"""

# ============================================
# 1. EXCEL - ALUNOS FALTANTES DIA 29
# ============================================
print("=" * 70)
print("1. EXCEL - ALUNOS FALTANTES NO DIA 29")
print("=" * 70)

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\user\presenca-ativa-inteligente\relatorios\Relatorio_Consolidado_BuscaAtiva.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
idx_29 = headers.index('29') + 1

faltantes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    valor_29 = row[idx_29 - 1]
    if valor_29 is not None and str(valor_29).strip() == '-':
        faltantes.append({
            'turma': row[0],
            'nome': row[2],
            'ra': row[3],
            'valor_29': valor_29
        })

print(f"Total de faltantes no dia 29: {len(faltantes)}")
print("\n--- 10 exemplos de faltantes ---")
for i, a in enumerate(faltantes[:10], 1):
    print(f"{i:2}. {a['nome']}")
    print(f"     RA: {a['ra']}")
    print(f"     Turma: {a['turma']}")
    print(f"     Valor coluna 29: '{a['valor_29']}'")

# ============================================
# 2. ESTADO DO SUPABASE
# ============================================
print("\n" + "=" * 70)
print("2. ESTADO DO SUPABASE (schema: busca_ativa_v2)")
print("=" * 70)

from app.infrastructure.supabase.repositories import SupabaseRepository
from datetime import datetime, timezone, date

repo = SupabaseRepository()
schema = 'busca_ativa_v2'
hoje = date(2026, 4, 29)

# HELPER: contar raw_inbound
all_raw = repo.client.schema(schema).table('raw_inbound').select('*').execute().data or []
pendentes = repo.list_unprocessed_raw_inbound(limit=1000)
raw_hoje = [r for r in all_raw if (r.get('received_at') or r.get('created_at') or '')[:10] == str(hoje)]

print(f"\n--- 2.1 raw_inbound ---")
print(f"Total: {len(all_raw)}")
print(f"Pendentes (processed=false): {len(pendentes)}")
print(f"Registros de hoje (29/04): {len(raw_hoje)}")
if raw_hoje:
    for r in raw_hoje[:3]:
        print(f"  ID: {r['id']}, processed: {r['processed']}, error: {str(r.get('processing_error',''))[:80]}")

# HELPER: responses
all_resp = repo.client.schema(schema).table('responses').select('*').execute().data or []
resp_hoje = [r for r in all_resp if (r.get('created_at') or '')[:10] == str(hoje)]
from collections import Counter
conf_counts = Counter(r.get('identity_confidence','?') for r in all_resp)

print(f"\n--- 2.2 responses ---")
print(f"Total: {len(all_resp)}")
print(f"De hoje (29/04): {len(resp_hoje)}")
print("Distribuicao identity_confidence (total):")
for conf, cnt in conf_counts.most_common():
    print(f"  '{conf}': {cnt}")

# HELPER: messages
all_msgs = repo.client.schema(schema).table('messages').select('*').execute().data or []
msgs_hoje = [m for m in all_msgs if (m.get('created_at') or '')[:10] == str(hoje)]

print(f"\n--- 2.3 messages ---")
print(f"Total: {len(all_msgs)}")
print(f"De hoje (29/04): {len(msgs_hoje)}")
if msgs_hoje:
    status_counter = Counter(m.get('status','?') for m in msgs_hoje)
    print("Status das mensagens de hoje:")
    for s,c in status_counter.items():
        print(f"  '{s}': {c}")

# HELPER: campaigns
all_camps = repo.client.schema(schema).table('campaigns').select('*').execute().data or []

print(f"\n--- 2.4 campaigns ---")
print(f"Total: {len(all_camps)}")
print("Campanhas existentes:")
for c in all_camps:
    print(f"  - Nome: '{c.get('name')}'")
    print(f"    absence_days: '{c.get('absence_days')}'")
    print(f"    status: '{c.get('status')}'")
    print(f"    dispatched_at: {c.get('dispatched_at')}")
    print()

# HELPER: students
all_students = repo.client.schema(schema).table('students').select('*').execute().data or []

print(f"\n--- 2.5 students vs Excel ---")
print(f"Students cadastrados no Supabase: {len(all_students)}")
print(f"Alunos na planilha Excel: {ws.max_row - 1}")
print(f"ATENCAO: Excel tem {ws.max_row - 1} alunos, Supabase tem apenas {len(all_students)} students")

# ============================================
# 3. DIAGNÓSTICO - INBOUND NÃO PROCESSADO
# ============================================
print("\n" + "=" * 70)
print("3. DIAGNÓSTICO - INBOUND NAO ESTÁ SENDO PROCESSADO")
print("=" * 70)

print(f"\n--- 3.1 Análise dos {len(pendentes)} pendentes ---")
if pendentes:
    tipos = Counter()
    for p in pendentes:
        payload = p.get('payload', {})
        msg_type = 'unknown'
        if isinstance(payload, dict):
            msg = payload.get('data', {}).get('message', {})
            if msg:
                for k in msg.keys():
                    msg_type = k
                    break
        tipos[msg_type] += 1
    
    print("Tipos de mensagens pendentes:")
    for t, cnt in tipos.most_common():
        print(f"  {t}: {cnt}")
    
    erros = Counter(str(p.get('processing_error',''))[:80] for p in pendentes)
    print("\nErros mais comuns:")
    for err, cnt in erros.most_common(3):
        print(f"  ({cnt}x) {err}")

print("\n--- 3.2 Possíveis causas da falha ---")
print("""
[CAUSA 1] TIMEOUT DE CONEXAO COM SUPABASE - CONFIRMADA
  - 79 registros com erro WinError 10060 (timeout de rede)
  - Aconteceu em lote durante processamento do dia 28/04
  - Supabase cliente Python usa httpx com timeout ~30s
  - Se Supabase lento ou rede instavel, ocorre timeout
  - EVIDENCIA: Teste manual hoje funciona (evento transitivo)

[CAUSA 2] IDENTIDADE NAO RESOLVIDA
  - Responses: 32/36 UNRESOLVED (guardian_id NULL, student_id NULL)
  - Students: apenas 33 cadastrados vs 265 no Excel
  - Sem student vinculado, nao gera campaign_message
  - Porem: nao causa timeout, apenas deixa resposta isolada

[CAUSA 3] WORKER PARADO OU NAO RODANDO EM LOOP
  - 79 mensagens aguardando desde 28/04
  - Se worker automatico nao esta em execucao continua, ficam pendentes
  - O worker ler e processará (exceto se timeout recorrer)
""")

# ============================================
# 4. SUGESTÃO DE CORREÇÃO IMEDIATA
# ============================================
print("\n" + "=" * 70)
print("4. SUGESTAO DE CORRECAO IMEDIATA")
print("=" * 70)

print("""
AÇÃO IMEDIATA 1: Executar reprocessamento manual
  $ python -m app.workers.reprocess_inbound --limit 80
  
  Se timeout persistir, aumentar timeout do Supabase client:
  → Modificar: app/infrastructure/supabase/repositories.py
  → Linha 25-27 (create_client), adicionar ClientOptions:
    
    from supabase.lib.client_options import ClientOptions
    options = ClientOptions(postgrest_client_timeout=120.0)
    self._client = create_client(settings.supabase_url, settings.supabase_key, options=options)

AÇÃO IMEDIATA 2: Sincronizar students com Excel
  - Supabase: 33 students | Excel: 265 alunos
  - Script para importar RA, nome, turma para students
  - Criar: scripts/sync_students_from_excel.py
  - Atencao: formatar RA corretamente (remover " /SP" se necessario)

AÇÃO IMEDIATA 3: Criar campanha ativa para 29/04/2026
  - Nenhuma campanha com absence_days='29/04/2026'
  - INSERT em campaigns com:
      name = 'Busca Ativa 29/04 - faltantes'
      absence_days = '29/04/2026'
      status = 'active' (ou 'scheduled' + dispatch)
      school_id = 'aac99735-32cb-4615-b2cb-0be315f18374'

AÇÃO IMEDIATA 4: Disparar follow-up após sync
  - Com students sincronizados e campanha ativa:
  $ python scripts/followup_campaign_v2.py --campaign-id <id>

ARQUIVO PARA MODIFICAR (timeout):
  app/infrastructure/supabase/repositories.py (linhas 25-27)

ARQUIVO PARA CRIAR (sync):
  scripts/sync_students_from_excel.py

RESUMO PRIORITARIO:
  1. Aumentar timeout Supabase (parar novos timeouts)
  2. Executar reprocess_inbound (limpar os 79 pendentes)
  3. Sincronizar students com Excel (ter base completa de 265)
  4. Criar campanha 29/04 e disparar follow-up
""")

print("\n" + "=" * 70)
print("RELATORIO CONCLUIDO")
print("=" * 70)
