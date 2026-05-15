"""
RELATORIO FINAL DE EXPLORACAO - Busca Ativa Inteligente
Data: 2026-04-29
"""

import openpyxl
from app.infrastructure.supabase.repositories import SupabaseRepository
from collections import Counter
from datetime import date

# ============================================
# SECAO 1: EXCEL - FALTANTES DIA 29
# ============================================
print("=" * 70)
print("SECAO 1: EXCEL - ALUNOS FALTANTES NO DIA 29")
print("=" * 70)

wb = openpyxl.load_workbook(r'C:\Users\user\presenca-ativa-inteligente\relatorios\Relatorio_Consolidado_BuscaAtiva.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
idx_29 = headers.index('29') + 1

faltantes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    valor_29 = row[idx_29 - 1]
    if valor_29 is not None and str(valor_29).strip() == '-':
        faltantes.append({
            'turma': row[0],
            'nome': row[2],
            'ra': row[3],
        })

print(f"Total de faltantes no dia 29: {len(faltantes)}")
print("\n--- 10 exemplos de faltantes ---")
for i, a in enumerate(faltantes[:10], 1):
    nome = a['nome']
    ra = a['ra']
    turma = a['turma'][:45]  # truncar para caber
    print(f"{i:2}. {nome}")
    print(f"     RA: {ra}")
    print(f"     Turma: {turma}")

# ============================================
# SECAO 2: ESTADO DO SUPABASE
# ============================================
print("\n" + "=" * 70)
print("SECAO 2: ESTADO DO SUPABASE (schema: busca_ativa_v2)")
print("=" * 70)

repo = SupabaseRepository()
schema = 'busca_ativa_v2'
hoje = date(2026, 4, 29)

# Helper para obter dados (com try/except para timeout)
def safe_query(fn, default=None):
    try:
        return fn()
    except Exception as e:
        print(f"  [TIMEOUT/CONNECT ERROR: {type(e).__name__}]")
        return default

print("\n--- 2.1 raw_inbound ---")
all_raw = safe_query(
    lambda: repo.client.schema(schema).table('raw_inbound').select('*').execute().data or [],
    default=[]
)
pendentes = safe_query(
    lambda: repo.list_unprocessed_raw_inbound(limit=1000),
    default=[]
)

raw_hoje = [r for r in all_raw if (r.get('received_at') or r.get('created_at') or '')[:10] == str(hoje)]

print(f"Total: {len(all_raw)}")
print(f"Pendentes (processed=false): {len(pendentes)}")
print(f"De hoje (29/04): {len(raw_hoje)}")
if raw_hoje:
    for r in raw_hoje[:2]:
        print(f"  ID: {r['id'][:8]}..., processed={r['processed']}, error={str(r.get('processing_error',''))[:60]}")

print("\n--- 2.2 responses ---")
all_resp = safe_query(
    lambda: repo.client.schema(schema).table('responses').select('*').execute().data or [],
    default=[]
)
resp_hoje = [r for r in all_resp if (r.get('created_at') or '')[:10] == str(hoje)]
conf_counts = Counter(r.get('identity_confidence','?') for r in all_resp)

print(f"Total: {len(all_resp)}")
print(f"De hoje (29/04): {len(resp_hoje)}")
print("Distribuicao identity_confidence:")
for conf, cnt in conf_counts.most_common():
    sem = 'SEM GUARDIAN' if conf == 'UNRESOLVED' else ''
    print(f"  '{conf}': {cnt} {sem}")

print("\n--- 2.3 messages ---")
all_msgs = safe_query(
    lambda: repo.client.schema(schema).table('messages').select('*').execute().data or [],
    default=[]
)
msgs_hoje = [m for m in all_msgs if (m.get('created_at') or '')[:10] == str(hoje)]

print(f"Total: {len(all_msgs)}")
print(f"De hoje (29/04): {len(msgs_hoje)}")
if msgs_hoje:
    st = Counter(m.get('status','?') for m in msgs_hoje)
    for s,c in st.items():
        print(f"  Status '{s}': {c}")
else:
    print("  Nenhuma message enviada hoje")

print("\n--- 2.4 campaigns ---")
all_camps = safe_query(
    lambda: repo.client.schema(schema).table('campaigns').select('*').execute().data or [],
    default=[]
)
print(f"Total: {len(all_camps)}")
print("Campanhas:")
for c in all_camps:
    nome = c.get('name','')[:50]
    d = c.get('absence_days','')
    st = c.get('status','')
    dp = c.get('dispatched_at')
    print(f"  - '{nome}'")
    print(f"    Dias: {d} | Status: {st} | Disparada: {dp}")

print("\n--- 2.5 students ---")
all_students = safe_query(
    lambda: repo.client.schema(schema).table('students').select('*').execute().data or [],
    default=[]
)
print(f"Students cadastrados no Supabase: {len(all_students)}")
print(f"Alunos na planilha Excel:        {ws.max_row - 1}")
print(f"GAP de sincronizacao: { (ws.max_row - 1) - len(all_students) } alunos faltando no Supabase")

# ============================================
# SECAO 3: DIAGNOSTICO
# ============================================
print("\n" + "=" * 70)
print("SECAO 3: DIAGNOSTICO - POR QUE INBOUND NAO PROCESSA?")
print("=" * 70)

print(f"\n--- 3.1 Analise dos {len(pendentes)} pendentes ---")
if pendentes:
    tipos = Counter()
    for p in pendentes:
        payload = p.get('payload', {})
        msg_type = 'unknown'
        if isinstance(payload, dict):
            msg = payload.get('data', {}).get('message', {})
            if msg:
                for k in msg.keys():
                    msg_type = k
                    break
        tipos[msg_type] += 1
    print("Tipos de mensagens pendentes:")
    for t, cnt in tipos.most_common():
        print(f"  {t}: {cnt}")

    erros = Counter(str(p.get('processing_error',''))[:70] for p in pendentes)
    print("\nErros mais comuns:")
    for err, cnt in erros.most_common(3):
        # Simplificar erro
        err_short = err.replace('[WinError 10060]', '[TIMEOUT 10060]')
        print(f"  ({cnt}x) {err_short}")

print("\n--- 3.2 Causas identificadas ---")
print("""
[CAUSA PRINCIPAL] Timeout de conexao com Supabase
  - Sintoma: 79 registros com erro ConnectTimeout/WinError 10060
  - Ocorreu em lote durante processamento do dia 28/04
  - Tipo de erro: ConnectTimeout (conexao TCP nao estabelecida a tempo)
  - Todas as 79 mensagens sao do mesmo periodo (28/04 17:52-18:12)
  - Isso indica instabilidade de rede ou saturacao do Supabase PostgREST
  - EVIDENCIA: Reprocessamento manual hoje funcionou (condicao normalizada)
  - Porém: novas consultas ainda esbarram em timeout intermitente

[CAUSA SECUNDARIA] Base de students desatualizada
  - Supabase tem apenas 33 students cadastrados
  - Planilha tem 265 alunos (240 faltantes no dia 29)
  - Sem student no Supabase, nao eh possivel vincular response a aluno
  - Consequencia: 32/36 responses ficam UNRESOLVED (guardian_id=NULL, student_id=NULL)
  - Isso nao causa timeout, mas compromete analytics e campanhas

[CAUSA TERCIARIA] Nenhuma campanha ativa para dia 29
  - Todas as 6 campanhas estao status 'draft'
  - Nenhuma campanha com absence_days='29/04/2026'
  - Sem campanha ativa, nao ha disparo automatico para os 240 faltantes
  - Isso e consequencia da base nao sincronizada, nao causa do timeout
""")

# ============================================
# SECAO 4: CORRECOES IMEDIATAS
# ============================================
print("\n" + "=" * 70)
print("SECAO 4: PLANO DE CORRECAO IMEDIATA")
print("=" * 70)

print("""
PASSO 1: AUMENTAR TIMEOUT DO CLIENTE SUPABASE
  Arquivo: app/infrastructure/supabase/repositories.py
  Linha:   ~27 (onde create_client e chamado)
  Alteracao:
    from supabase.lib.client_options import ClientOptions
    options = ClientOptions(postgrest_client_timeout=120.0)
    self._client = create_client(settings.supabase_url, settings.supabase_key, options=options)
  
  Justificativa: Evita ConnectTimeout em redes instaveis ou Supabase lento

PASSO 2: EXECUTAR REPROCESSAMENTO DOS 79 PENDENTES
  Comando: python -m app.workers.reprocess_inbound --limit 80
  Isso deve limpar os registros de 28/04 que falharam com timeout
  Se ainda falhar, rodar novamente (o erro era intermitente)

PASSO 3: SINCRONIZAR STUDENTS COM EXCEL
  Arquivo a criar: scripts/sync_students_from_excel.py
  Logica:
    - Ler planilha Consolidado (265 linhas)
    - Para cada linha: extrair Nome, RA, Turma
    - Normalizar RA (remover ' /SP' se houver)
    - UPSERT na tabela students (school_id fixo)
  Resultado esperado: 265 students no Supabase

PASSO 4: CRIAR CAMPANHA PARA 29/04/2026
  Inserir em campaigns:
    name:           'Busca Ativa 29/04 - 240 faltantes'
    absence_days:  '29/04/2026'
    school_id:     'aac99735-32cb-4615-b2cb-0be315f18374'
    status:        'draft'  (ou 'active' se ja for disparar)
    created_at:    now()
  Observacao: Na ausencia de script de criacao, usar Supabase UI ou SQL direto

PASSO 5: DISPARAR FOLLOW-UP
  Comando: python scripts/followup_campaign_v2.py --campaign-id <id_novo>
  Isso enviara mensagens aos 240 responsaveis (se guardians estiverem vinculados)

PASSO 6: VERIFICAR VINCULO STUDENT_GUARDIANS
  Opcional: Verificar se a tabela student_guardians tem guardians dos 240 alunos
  Se nao houver, pode ser necessario importar tambem
""")

print("\n" + "=" * 70)
print("RELATORIO CONCLUIDO")
print("=" * 70)
print("\nArquivos modificados/criados sugeridos:")
print("  MODIFICAR:  app/infrastructure/supabase/repositories.py (timeout)")
print("  CRIAR:      scripts/sync_students_from_excel.py")
print("  CRIAR:      scripts/create_campaign_29_04.sql (SQL manual)")
print("\nComandos a executar:")
print("  1. python -m app.workers.reprocess_inbound --limit 80")
print("  2. python scripts/sync_students_from_excel.py")
print("  3. python scripts/followup_campaign_v2.py")
print("=" * 70)
