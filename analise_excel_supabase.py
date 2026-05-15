import openpyxl
import requests
import json
from datetime import datetime, date

# ============================================
# 1. PROCESSAR EXCEL
# ============================================
print("=" * 60)
print("1. ANALISANDO EXCEL - Relatorio_Consolidado_BuscaAtiva.xlsx")
print("=" * 60)

wb = openpyxl.load_workbook('relatorios/Relatorio_Consolidado_BuscaAtiva.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

# Encontrar índice da coluna '29'
idx_29 = headers.index('29') + 1  # openpyxl é 1-based
print(f"Coluna '29' encontrada no índice: {idx_29}")
print(f"Colunas: {headers[:10]}...")

alunos_faltantes_dia29 = []
alunos_presentes_dia29 = []

for row in ws.iter_rows(min_row=2, values_only=True):
    turma = row[0]
    num = row[1]
    nome = row[2]
    ra = row[3]
    valor_29 = row[idx_29 - 1]  # ajuste 0-based

    if valor_29 is None:
        continue

    valor_str = str(valor_29).strip()
    # Regra: '-' = falta, qualquer número = presença (1, 6, 7, 8, 9, 2, etc)
    if valor_str == '-':
        alunos_faltantes_dia29.append({
            'turma': turma,
            'nome': nome,
            'ra': ra,
            'valor_29': valor_29
        })
    else:
        alunos_presentes_dia29.append({
            'turma': turma,
            'nome': nome,
            'ra': ra,
            'valor_29': valor_29
        })

print(f"\nTotal de alunos na planilha: {ws.max_row - 1}")
print(f"Alunos presentes no dia 29: {len(alunos_presentes_dia29)}")
print(f"Alunos faltantes no dia 29: {len(alunos_faltantes_dia29)}")

print("\n--- EXEMPLOS DE FALTANTES NO DIA 29 (máximo 10) ---")
for i, aluno in enumerate(alunos_faltantes_dia29[:10], 1):
    print(f"{i}. {aluno['nome']} | RA: {aluno['ra']} | Turma: {aluno['turma']} | Valor: '{aluno['valor_29']}'")

# ============================================
# 2. CONSULTAR SUPABASE
# ============================================
print("\n" + "=" * 60)
print("2. CONSULTANDO SUPABASE")
print("=" * 60)

SUPABASE_URL = "https://cpniwvghxlkposaeyboa.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNwbml3dmdoeGxrcG9zYWV5Ym9hIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTI5MDc3MiwiZXhwIjoyMDkwODY2NzcyfQ.aGCEqHjue7sORZQ_Wa0OVS5fzSIKkS08OxV3ewb7HP8"

headers_supabase = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

# Hoje é 2026-04-29
today = "2026-04-29"
print(f"Data de hoje: {today}")

# 2.1 Verificar raw_inbound de hoje
try:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/raw_inbound",
        headers=headers_supabase,
        params={
            "select": "count",
            "created_at": f"gte.{today}T00:00:00",
            "created_at": f"lte.{today}T23:59:59"
        }
    )
    if resp.status_code == 200:
        count_raw = len(resp.json()) if isinstance(resp.json(), list) else 0
        print(f"[OK] raw_inbound hoje: {count_raw} registros")
    else:
        print(f"[ERRO] raw_inbound status {resp.status_code}: {resp.text[:200]}")
except Exception as e:
    print(f"[ERRO] raw_inbound: {e}")

# 2.2 Verificar responses de hoje
try:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/responses",
        headers=headers_supabase,
        params={
            "select": "count",
            "created_at": f"gte.{today}T00:00:00",
            "created_at": f"lte.{today}T23:59:59"
        }
    )
    if resp.status_code == 200:
        count_resp = len(resp.json()) if isinstance(resp.json(), list) else 0
        print(f"[OK] responses hoje: {count_resp} registros")
    else:
        print(f"[ERRO] responses status {resp.status_code}: {resp.text[:200]}")
except Exception as e:
    print(f"[ERRO] responses: {e}")

# 2.3 Verificar campanha ativa (campaigns com absence_days contendo '29/04/2026' ou '2026-04-29')
try:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/campaigns",
        headers=headers_supabase,
        params={"select": "*"}
    )
    if resp.status_code == 200:
        campaigns = resp.json()
        print(f"[OK] Total campaigns: {len(campaigns)}")
        for camp in campaigns:
            print(f"   - ID: {camp.get('id')}, absence_days: {camp.get('absence_days')}, status: {camp.get('status')}")
    else:
        print(f"[ERRO] campaigns status {resp.status_code}")
except Exception as e:
    print(f"[ERRO] campaigns: {e}")

# 2.4 Verificar mensagens enviadas hoje
try:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/messages",
        headers=headers_supabase,
        params={
            "select": "id,status,created_at",
            "created_at": f"gte.{today}T00:00:00",
            "created_at": f"lte.{today}T23:59:59"
        }
    )
    if resp.status_code == 200:
        messages = resp.json()
        print(f"[OK] messages enviadas hoje: {len(messages)}")
        status_count = {}
        for msg in messages:
            s = msg.get('status', 'unknown')
            status_count[s] = status_count.get(s, 0) + 1
        for s, c in status_count.items():
            print(f"   - Status '{s}': {c} mensagens")
    else:
        print(f"[ERRO] messages status {resp.status_code}")
except Exception as e:
    print(f"[ERRO] messages: {e}")

# 2.5 Verificar raw_inbound não processados (processed = false)
try:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/raw_inbound",
        headers=headers_supabase,
        params={
            "select": "id,processed,error,created_at",
            "processed": "eq.0"
        }
    )
    if resp.status_code == 200:
        inbound_pending = resp.json()
        print(f"\n[OK] raw_inbound NÃO PROCESSADOS: {len(inbound_pending)}")
        if inbound_pending:
            print("   Primeiros 5 registros:")
            for rec in inbound_pending[:5]:
                print(f"   - ID: {rec.get('id')}, Error: {rec.get('error')}, Created: {rec.get('created_at')}")
    else:
        print(f"[ERRO] raw_inbound pending status {resp.status_code}")
except Exception as e:
    print(f"[ERRO] raw_inbound pending: {e}")

print("\n" + "=" * 60)
print("CONCLUSÃO - DADOS EXTRAÍDOS")
print("=" * 60)
print(f"Faltantes no dia 29: {len(alunos_faltantes_dia29)} alunos")
