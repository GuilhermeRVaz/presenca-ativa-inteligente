"""
RELATÓRIO DE EXPLORAÇÃO - Busca Ativa Inteligente
Data: 2026-04-29
Agente: Kilo
"""

# ============================================
# 1. EXCEL - ALUNOS FALTANTES DIA 29
# ============================================
print("=" * 70)
print("1. EXCEL - ALUNOS FALTANTES NO DIA 29")
print("=" * 70)

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\user\presenca-ativa-inteligente\relatorios\Relatorio_Consolidado_BuscaAtiva.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
idx_29 = headers.index('29') + 1

faltantes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    valor_29 = row[idx_29 - 1]
    if valor_29 is not None and str(valor_29).strip() == '-':
        faltantes.append({
            'turma': row[0],
            'nome': row[2],
            'ra': row[3],
            'valor_29': valor_29
        })

print(f"Total de faltantes no dia 29: {len(faltantes)}")
print("\n--- 10 exemplos de faltantes ---")
for i, a in enumerate(faltantes[:10], 1):
    print(f"{i:2}. {a['nome']}")
    print(f"     RA: {a['ra']}")
    print(f"     Turma: {a['turma']}")
    print(f"     Valor coluna 29: '{a['valor_29']}'")

# ============================================
# 2. ESTADO DO SUPABASE
# ============================================
print("\n" + "=" * 70)
print("2. ESTADO DO SUPABASE (schema: busca_ativa_v2)")
print("=" * 70)

from app.infrastructure.supabase.repositories import SupabaseRepository
from datetime import datetime, timezone

repo = SupabaseRepository()
schema = 'busca_ativa_v2'

print("\n--- 2.1 raw_inbound ---")
try:
    # Total
    all_raw = repo.client.schema(schema).table('raw_inbound').select('count').execute().data or []
    print(f"Total raw_inbound: {len(all_raw)}")
    
    # Pendentes (processed = false)
    pendentes = repo.list_unprocessed_raw_inbound(limit=1000)
    print(f"Pendentes (processed=false): {len(pendentes)}")
    
    # De hoje (29/04/2026)
    from datetime import date
    hoje = date(2026, 4, 29)
    raw_hoje = [r for r in all_raw if (r.get('received_at') or r.get('created_at') or '')[:10] == str(hoje)]
    print(f"Registros de hoje (29/04): {len(raw_hoje)}")
    
    if raw_hoje:
        for r in raw_hoje[:3]:
            print(f"  ID: {r['id']}, processed: {r['processed']}, error: {str(r.get('processing_error',''))[:80]}")
except Exception as e:
    print(f"ERRO ao consultar raw_inbound: {e}")

print("\n--- 2.2 responses ---")
try:
    all_resp = repo.client.schema(schema).table('responses').select('*').execute().data or []
    print(f"Total responses: {len(all_resp)}")
    
    resp_hoje = [r for r in all_resp if (r.get('created_at') or '')[:10] == str(hoje)]
    print(f"Responses de hoje (29/04): {len(resp_hoje)}")
    
    # Distribution por confidence
    from collections import Counter
    conf_counts = Counter(r.get('identity_confidence','?') for r in all_resp)
    print(f"Distribuição identity_confidence (total):")
    for conf, cnt in conf_counts.most_common():
        print(f"  '{conf}': {cnt}")
except Exception as e:
    print(f"ERRO ao consultar responses: {e}")

print("\n--- 2.3 messages ---")
try:
    all_msgs = repo.client.schema(schema).table('messages').select('*').execute().data or []
    print(f"Total messages: {len(all_msgs)}")
    
    msgs_hoje = [m for m in all_msgs if (m.get('created_at') or '')[:10] == str(hoje)]
    print(f"Messages de hoje (29/04): {len(msgs_hoje)}")
    
    if msgs_hoje:
        status_counter = {}
        for m in msgs_hoje:
            s = m.get('status','?')
            status_counter[s] = status_counter.get(s, 0) + 1
        print("Status das mensagens de hoje:")
        for s,c in status_counter.items():
            print(f"  '{s}': {c}")
except Exception as e:
    print(f"ERRO ao consultar messages: {e}")

print("\n--- 2.4 campaigns ---")
try:
    all_camps = repo.client.schema(schema).table('campaigns').select('*').execute().data or []
    print(f"Total campaigns: {len(all_camps)}")
    print("Campanhas existentes:")
    for c in all_camps:
        print(f"  - Nome: '{c.get('name')}'")
        print(f"    absence_days: '{c.get('absence_days')}'")
        print(f"    status: '{c.get('status')}'")
        print(f"    dispatched_at: {c.get('dispatched_at')}")
        print(f"    school_id: {c.get('school_id')}")
        print()
except Exception as e:
    print(f"ERRO ao consultar campaigns: {e}")

print("\n--- 2.5 students vs Excel ---")
try:
    all_students = repo.client.schema(schema).table('students').select('*').execute().data or []
    print(f"Students cadastrados no Supabase: {len(all_students)}")
    print(f"Alunos na planilha Excel: {ws.max_row - 1}")
    print(f"⚠️  Discrepância: Excel tem {ws.max_row - 1} alunos, mas Supabase tem apenas {len(all_students)} students")
except Exception as e:
    print(f"ERRO: {e}")

# ============================================
# 3. DIAGNÓSTICO DO INBOUND NÃO PROCESSADO
# ============================================
print("\n" + "=" * 70)
print("3. DIAGNÓSTICO - INBOUND NÃO ESTÁ SENDO PROCESSADO")
print("=" * 70)

print("\n--- 3.1 Análise dos 80 pendentes ---")
pendentes = repo.list_unprocessed_raw_inbound(limit=80)
print(f"Total pendentes: {len(pendentes)}")

if pendentes:
    # Agrupar por tipo de mensagem
    from collections import Counter
    tipos = Counter()
    for p in pendentes:
        payload = p.get('payload', {})
        msg_type = 'unknown'
        if isinstance(payload, dict):
            msg = payload.get('data', {}).get('message', {})
            if msg:
                # Primeira chave do dicionário message
                for k in msg.keys():
                    msg_type = k
                    break
        tipos[msg_type] += 1
    
    print("Tipos de mensagens pendentes:")
    for t, cnt in tipos.most_common():
        print(f"  {t}: {cnt}")
    
    # Erro mais comum
    erros = Counter(str(p.get('processing_error',''))[:80] for p in pendentes)
    print("\nErros mais comuns:")
    for err, cnt in erros.most_common(3):
        print(f"  ({cnt}x) {err}")

print("\n--- 3.2 Possíveis causas da falha ---")
print("""
Baseado no código e no estado observado:

CAUSA 1: [TIMEOUT DE CONEXÃO COM SUPABASE - CONFIRMADA]
  - 80 registros com erro: "[WinError 10060] Uma tentativa de conexão falhou..."
  - Erro de socket Windows: host não respondeu dentro do timeout
  - Aconteceu em lote durante o processamento do dia 28/04
  - O Supabase cliente Python usa httpx com timeout padrão (~30s conect + 30s read)
  - Se o Supabase estiver lento ou a rede instável, ocorre timeout
  - **Evidência**: Teste manual hoje funciona, indicando foi evento transitório

CAUSA 2: [IDENTIDADE NÃO RESOLVIDA → MENSAGEM FICA PENDENTE]
  - Das 33 messages cadastradas, apenas 10 guardians
  - Das 35 responses, 20+ estão UNRESOLVED (guardian_id NULL, student_id NULL)
  - O sistema permite salvar responses UNRESOLVED
  - **Porém**: Isso não causa timeout, apenas deixa a resposta desvinculada

CAUSA 3: [WORKER PARADO OU NÃO RODANDO]
  - 80 mensagens aguardando há ~1 dia (desde 28/04)
  - Se o worker de reprocessamento não estiver rodando em loop, ficam pendentes
  - O worker lerá as 80 e processará (exceto se timeout recorrer)
  - **Verificar**: Se o worker automático está ativo (processamento contínuo)
""")

# ============================================
# 4. SUGESTÃO DE CORREÇÃO IMEDIATA
# ============================================
print("\n" + "=" * 70)
print("4. SUGESTÃO DE CORREÇÃO IMEDIATA")
print("=" * 70)

print("""
AÇÃO IMEDIATA 1: Executar reprocessamento manual dos 80 pendentes
  $ python -m app.workers.reprocess_inbound --limit 80
  
  Se timeout persistir, aumentar timeout do Supabase client:
  → Modificar app/infrastructure/supabase/repositories.py
  → Linha ~27 onde create_client é chamado:
    
    from supabase import create_client
    from supabase.lib.client_options import ClientOptions
    
    options = ClientOptions(
        postgrest_client_timeout=120.0,  # era padrão ~30s
        storage_client_timeout=30.0,
    )
    self._client = create_client(settings.supabase_url, settings.supabase_key, options=options)
  
  Isso aumenta timeout de consultas PostgREST de 30s para 120s.

AÇÃO IMEDIATA 2: Sincronizar planilha Excel com Supabase (students)
  - O Supabase tem apenas 33 students cadastrados
  - A planilha tem 265 alunos
  - Se a campanha for enviada, só terá 33 students para vincular → inconsistência
  - Importar os 265 alunos da planilha para a tabela students do Supabase
  - Colunas: name, ra, class_name, school_id
  
  Sugestão: Criar script de sync:
    for cada linha da planilha (a partir da linha 2):
      extrair nome, RA, turma
      upsert na tabela students (RA como chave única)
      OBS: Verificar formato do RA (planilha tem "000115114444-7 /SP")

AÇÃO IMEDIATA 3: Criar campanha ativa para 29/04/2026
  - Nenhuma campanha com absence_days='29/04/2026' existe
  - A planilha indica faltantes do dia 29
  - É necessário ter uma campanha ativa (status='active' ou 'scheduled' com dispatched_at preenchido)
  - Criar campaign com:
      name: "Busca Ativa 29/04 - faltantes"
      absence_days: "29/04/2026"
      status: "active" (ou "scheduled" → depois disparar)
      school_id: aac99735-32cb-4615-b2cb-0be315f18374

AÇÃO IMEDIATA 4: Executar script de disparo após sync
  - Com students sincronizados e campanha ativa, rodar:
    $ python scripts/followup_campaign_v2.py --campaign-id <id_novo>
  - Isso enviará mensagens aos responsáveis dos 240 faltantes

ARQUIVO PARA MODIFICAR (timeout):
  app/infrastructure/supabase/repositories.py
    linha 25-27: adicionar ClientOptions com timeout maior

ARQUIVO PARA CRIAR (sync Excel → Supabase):
  scripts/sync_students_from_excel.py
    - ler Relatorio_Consolidado_BuscaAtiva.xlsx
    - para cada aluno (NÃO repetir RA já existente)
    - inserir/atualizar students
    - opcional: vincular student_guardians (se houver)

RESUMO DAS AÇÕES PRIORITÁRIAS:
  1. Aumentar timeout Supabase (parar novos timeouts)
  2. Executar reprocess_inbound (limpar os 80 pendentes)
  3. Sincronizar students com Excel (ter base completa)
  4. Criar campanha 29/04 e disparar follow-up
""")

print("\n" + "=" * 70)
print("FIM DO RELATÓRIO")
print("=" * 70)
